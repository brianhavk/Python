import numpy as np
import openpyxl
import matplotlib.pyplot as plt
import pygimli as pg
from pygimli.utils import rndig
import pybert as pb


# some of this should rather go to data
def getReciprocals(data, change=False, remove=False):
    """Compute data reciprocity from forward and backward data.

    The reciprocity (difference between forward and backward array divided by
    their mean) is computed and saved under the dataContainer field 'rec'

    Parameters
    ==========
    data : pg.DataContainerERT
        input data container to be changed inplace
    change : bool [True]
        compute current-weighted mean of forward and backward values
    remove : bool [False]
        remove backward data that are present as forward data
    """
    if not data.allNonZero('r'):
        data.set('r', data('u') / data('i'))

    unF = uniqueERTIndex(data)
    unB = uniqueERTIndex(data, reverse=True)
    rF, rB = [], []
    rec = np.zeros(data.size())
    data.set('rec', pg.Vector(data.size()))
    for iB in range(data.size()):
        if unB[iB] in unF:
            iF = int(np.nonzero(unF == unB[iB])[0][0])
            rF.append(data('r')[iF])
            rB.append(data('r')[iB])
            rec[iB] = (rF[-1]-rB[-1]) / (rF[-1]+rB[-1]) * 2
            data('rec')[iF] = rec[iB]
            IF, IB = data('i')[iF], data('i')[iB]  # use currents for weighting
            if change and data('valid')[iF]:
                data('r')[iF] = (rF[-1] * IF + rB[-1] * IB) / (IF + IB)
                data('i')[iF] = (IF**2 + IB**2) / (IF + IB)  # according weight
                data('u')[iF] = data('r')[iF] * data('i')[iF]
                if remove:
                    data('valid')[iB] = 0  # for adding all others later on

    print(len(rF), "reciprocals")
    if remove:
        data.removeInvalid()


def turnDDInside(data):
    """Turn dipole-dipole data so that A and M are inside and k>0."""
    # rearrange arrays as B-A--M-N = fwd or N-M--A-B =bwd (A-M in the middle)
    ab1 = np.minimum(data('a'), data('b'))  # left AB
    ab2 = np.maximum(data('a'), data('b'))  # right AB
    mn1 = np.minimum(data('m'), data('n'))  # left MN
    mn2 = np.maximum(data('m'), data('n'))  # right MN
    aa = ab2.copy()  # for B-A--M-N: A right
    bb = ab1.copy()  # B left
    mm = mn1.copy()  # M left
    nn = mn2.copy()  # N right
    iBwd = np.array(data('a')) > np.array(data('m'))  # N-M--A-B
    aa[iBwd] = ab1[iBwd]  # A left
    bb[iBwd] = ab2[iBwd]  # B right
    mm[iBwd] = mn2[iBwd]  # M right
    nn[iBwd] = mn1[iBwd]  # N left
    data.set('a', aa)
    data.set('b', bb)
    data.set('m', mm)
    data.set('n', nn)


def fillup(data):
    """Fill up upper right triangle with lower left."""
    turnDDInside(data)
    iFwd = uniqueERTIndex(data)
    iBwd = uniqueERTIndex(data, reverse=True)
    for i in range(data.size()):
        if data('m')[i] > data('a')[i]:
            if iBwd[i] in iFwd:
                data.markInvalid(i)
            else:
                dummy = data('a')[i]
                data('a')[i] = data('m')[i]
                data('m')[i] = dummy
                dummy = data('b')[i]
                data('b')[i] = data('n')[i]
                data('n')[i] = dummy

    data.removeInvalid()


def makeDataContainer(ABMN, filename=None, a=125, nel=0, pos=None):
    data = pg.DataContainerERT()
    if pos is not None:
        for posi in pos:
            data.createSensor(pg.Pos(posi[0], posi[1], 0))
    else:
        if nel == 0:
            nel = int(np.max(ABMN[:, :4]))
        for i in range(nel):
            data.createSensor(pg.Pos(a*i, 0, 0))

    data.resize(ABMN.shape[0])
    for i, key in enumerate(['a', 'b', 'n', 'm', 'u', 'i', 'err']):
        data.set(key, ABMN[:, i]-(key in ['a', 'b', 'm', 'n'])*1)
#        data.set(key, ABMN[:, i] - 1)

    data.set('k', pg.core.geometricFactors(data, dim=3))
    data.set('r', pg.abs(data('u')/data('i')))
    data.set('rhoa', pg.abs(data('r') * data('k')))
    data.markValid(data('u') > 0)
    data.checkDataValidity()
    if filename is not None:
        data.save(filename, 'a b m n u i err k rhoa')
    return data


def combineBigSmall(bigData, smallData):
    """Combine data with two currents by filling up."""
    uBig = uniqueERTIndex(bigData)
    uSma = uniqueERTIndex(smallData)
    data = pg.DataContainerERT(smallData)  # make copy not destroying
    for i, u in enumerate(uSma):
        if u in uBig:
            data.markInvalid(i)

    data.removeInvalid()
    data.add(bigData)
    return data


def readHSGdata(csvfile, loggerfile, emap=None, simple=True):
    """Read large-scale ERT data from LIAG timeseries processing result (csv).

    Parameters
    ----------
    csvfile : str
        Comma-separated file containing data columns
    loggerfile : str
        logger file mapping the channels of the loggers to electrodes
    emap : dict, optional
        dictionary mapping names to numbers (e.g. '9A' to 9 and '9B' to 10)
    simple : bool [True]
        use columns 7 to 10 instead of 7,9,11,13 for error values

    Returns
    -------
    DATA : np.ndarray
        data containing A, B, M, N, voltage, current and error as columns
    """
    if emap is None:
        emap = {str(i): i for i in range(100)}
    logger = np.genfromtxt(loggerfile, dtype=str, skip_header=1, delimiter=';')
    # %%
    log2pos = {}
    for logi in logger:
        log2pos[logi[0]] = np.reshape([emap[ii] for ii in logi[1:]], (3, 2))
    # %% read CSVfile
    abst = np.genfromtxt(csvfile, dtype=str, usecols=0, skip_header=1)
    # %%
    sss = [ss.replace('\\', '_').split('_') for ss in abst]
    print(sss[0])
    AA = [emap[ss[-7]] for ss in sss]  # A
    BB = [emap[ss[-6]] for ss in sss]
    LL = [ss[-5][0] for ss in sss]
#    AA = [int(ss.split('_')[3]) for ss in abst]  # A
#    BB = [int(ss.replace('\\', '_').split('_')[4]) for ss in abst]
#    LL = [ss.replace('\\', '_').split('_')[5][0] for ss in abst]
    # %% Spannungsdatei
    cols = (1, 2, 3, 7, 9, 11, 13)
    if simple:
        cols = (1, 2, 3, 7, 8, 9, 10)
    V1, V2, V3, E1, E2, E3, II = np.genfromtxt(csvfile, skip_header=1,
                                               delimiter='\t',
                                               unpack=True, usecols=cols)
    # %% generate arrays (exchange A and B such that V is always positive)
    A = np.repeat(np.minimum(AA, BB), 3)
    B = np.repeat(np.maximum(AA, BB), 3)
    # %%
    MM = np.array([log2pos[L][:, 0] for L in LL])
    NN = np.array([log2pos[L][:, 1] for L in LL])
    M = np.minimum(MM, NN).ravel()
    N = np.maximum(MM, NN).ravel()
    current = np.repeat(II, 3).ravel()
    V = np.column_stack((V1, V2, V3)).ravel() * 1e-3
    MSE = np.column_stack((E1, E2, E3)).ravel()
    if 'FFT' in csvfile.upper():
        E = 1 / MSE
    elif 'stack' in csvfile.lower():  # lock in relation
        E = np.abs(1 - MSE / 100)
    else:  # Lock-In
        a, b = 0.0271, 0.2949
        SNRdB = - np.log(MSE/a+1e-10) / b
        # E = 10**(-SNRdB/10) * 0.01
        E = SNRdB
    # %% stack together everything
    ALL = np.column_stack((A, B, M, N, V, current, E))
    VALID = ALL[np.isfinite(ALL[:, 4]), :]  # nonzero voltage
    return VALID


def uniqueERTIndex(data, nI=0, reverse=False):
    """Generate unique index from sensor indices A/B/M/N for matching

    Parameters
    ----------
    data : DataContainerERT
        data container holding a b m n field registered as indices (int)
    I : int [0]
        index to generate (multiply), by default (0) sensorCount
        if two data files with different sensorCount are compared make sure
        to use the same I for both
    reverse : bool [False]
        exchange current (A, B) with potential (M, N) for reciprocal analysis
    """
    if nI == 0:
        nI = data.sensorCount() + 1
    normABMN = {'a': np.minimum(data('a'), data('b')) + 1,
                'b': np.maximum(data('a'), data('b')) + 1,
                'm': np.minimum(data('m'), data('n')) + 1,
                'n': np.maximum(data('m'), data('n')) + 1}
    abmn = ['a', 'b', 'm', 'n']   # 1 2 8 7
    if reverse:
        abmn = ['m', 'n', 'a', 'b']   # 7 8 2 1
#        abmn = ['n', 'm', 'b', 'a']   # 7 8 2 1
    ind = 0
    for el in abmn:
        ind = ind * nI + normABMN[el]  # data(el)

    return np.array(ind, dtype=np.int64)


def deleteDuplicates(data, remove=True):
    """Search for multiple data and delete the one with lower current."""
    ind = uniqueERTIndex(data)
    num = 0
    for ui in np.unique(ind):
        uind = np.nonzero(ind == ui)[0]
        if len(uind) > 1:
            current = np.array(data('i')[uind])
            for u in uind[current < max(current)]:
                data.markInvalid(u)
                num += 1

    print("Found {:d} duplicates".format(num))
    if remove:
        data.removeInvalid()


def extractReciprocals(fwd, bwd):
    nMax = max(fwd.sensorCount(), bwd.sensorCount())
    unF = uniqueERTIndex(fwd, nI=nMax)
    unB = uniqueERTIndex(bwd, nI=nMax, reverse=True)
    rF, rB = [], []
    rec = np.zeros(bwd.size())
    both = pg.DataContainerERT(fwd)
    both.set('rec', pg.Vector(both.size()))
    back = pg.DataContainerERT(bwd)
    back.set('rec', pg.Vector(back.size()))
    for iB in range(bwd.size()):
        if unB[iB] in unF:
            iF = int(np.nonzero(unF == unB[iB])[0][0])
            rF.append(fwd('r')[iF])
            rB.append(bwd('r')[iB])
            rec[iB] = (rF[-1]-rB[-1]) / (rF[-1]+rB[-1]) * 2
            both('rec')[iF] = rec[iB]
            IF, IB = fwd('i')[iF], bwd('i')[iB]  # use currents for weighting
            both('r')[iF] = (rF[-1] * IF + rB[-1] * IB) / (IF + IB)
            both('i')[iF] = (IF**2 + IB**2) / (IF + IB)  # according to weight
            both('u')[iF] = fwd('r')[iF] * fwd('i')[iF]
            back('valid')[iB] = 0  # for adding all others later on
    print(len(rF), "reciprocals")
    back.removeInvalid()
    both.add(back)
    return rec, both


def showDDMatrix(MAT, fxm, cMap='jet', cMin=None, cMax=None, ax=None,
                 **kwargs):
    """Show matrix."""
    if ax is None:
        fig, ax = plt.subplots(figsize=(14, 12))

    im = ax.matshow(MAT)
    im.set_cmap(cMap)

    if cMin is not None and cMax is not None:
        im.set_clim(cMin, cMax)

    ax.grid(True)
    cb = None
    if kwargs.pop('colorBar', True):
        cb = ax.figure.colorbar(im, ticks=np.linspace(cMin, cMax, 7), ax=ax,
                                orientation=kwargs.pop('orientation',
                                                       'vertical'))
    ixm = {fxm[k]: k for k in fxm}
    ax.xaxis.set_ticks_position('both')
    ax.yaxis.set_ticks_position('both')
    xt = np.arange(0, len(MAT), 1)
    ax.set_xticks(xt)
    if kwargs.pop('oldStyle', False):
        xtl = [str(int(ixm[xti]))+'/'+str(int(ixm[xti])+1) for xti in xt]
    else:
        xtl = [str(int(ixm[xx]//100))+'/'+str(int(ixm[xx] % 100)) for xx in xt]
    ax.set_xticklabels(xtl, rotation='vertical')
    ax.set_yticks(xt)
    ax.set_yticklabels(xtl)
    ax.xaxis.set_label_position('top')
    ax.set_xlabel('A/B electrodes')
    ax.set_ylabel('M/N electrodes')
    return ax, cb


def showDDData(data, field='rhoa', **kwargs):
    """Show dipole-dipole data as a matrix."""
    ab = (np.minimum(data('a'), data('b')) + 1) * 100 + \
        np.maximum(data('a'), data('b')) + 1
    mn = (np.minimum(data('m'), data('n')) + 1) * 100 + \
        np.maximum(data('m'), data('n')) + 1

    vals = kwargs.pop('vals', data(field))
    cMin = kwargs.pop('cMin', min(vals))
    cMax = kwargs.pop('cMax', max(vals))
    cMap = kwargs.pop('cMap', 'jet')
    MAT, xm, _ = pg.viewer.mpl.generateMatrix(ab, mn, vals,
                                             full=kwargs.pop('full', True))
    MAT[MAT == 0] = np.NaN
    logScale = kwargs.pop('logScale', True)
    if logScale:
        MAT = np.log10(MAT)
        cMin = np.log10(cMin)
        cMax = np.log10(cMax)

    ax, cb = showDDMatrix(MAT, xm, cMin=cMin, cMax=cMax, cMap=cMap, **kwargs)
    if logScale:
        ti = np.linspace(0, 1, 7)
        # yt = cb.ax.get_yticks()
        # cb.ax.set_yticks(ti)
        ci = 10**((cMax-cMin) * ti + cMin)
        if cb:
            if kwargs.pop('orientation', 'vertical') == 'vertical':
                cb.ax.set_yticklabels(["%g" % rndig(ii, 2) for ii in ci])
            else:
                cb.ax.set_xticklabels(["%g" % rndig(ii, 2) for ii in ci])

    if 'mark' in kwargs:
        # uu = pg.unique(pg.cat(ab, mn))
        for i, m in enumerate(kwargs['mark']):
            if m > 0:
                ax.plot(xm[ab[i]], xm[mn[i]], 'wx', markersize=4)
#                iab = pg.find(uu == ab[i])[0]
#                imn = pg.find(uu == mn[i])[0]
#                print(i, ab[i], mn[i], iab, imn, data('rec')[i])
#                print(xm[ab[i]], xm[mn[i]])
#                ax.plot(iab, imn, 'kx', markersize=2)

    if 'hlines' in kwargs:
        ax.hlines(kwargs['hlines'], *(ax.get_xlim()))
    if 'vlines' in kwargs:
        ax.vlines(kwargs['vlines'], *(ax.get_ylim()))
    ax.plot(ax.get_xlim(), ax.get_ylim()[::-1], 'k-')
    return ax


def showDDDataOld(data, field='rhoa', **kwargs):
    ab = np.minimum(data('a'), data('b')) + 1
    mn = np.minimum(data('m'), data('n')) + 1
    if isinstance(field, str):
        field = data(field)

    return pg.viewer.mpl.plotVecMatrix(ab, mn, field, **kwargs)


def readUniLdata(xlsxfile, pos, minC=6, maxC=49):
    """Read Excel sheet containing data from Uni Leipzig."""
    wb = openpyxl.load_workbook(xlsxfile)
    WB = wb.get_sheet_by_name(wb.sheetnames[0])  # big
    WS = wb.get_sheet_by_name(wb.sheetnames[1])  # small
    ra = range(minC, maxC)
    A = [int(WB.cell(row=4, column=i).value.split('-')[0])-1 for i in ra]
    current = [WB.cell(row=6, column=i).value for i in ra]
    cSmall = [WS.cell(row=6, column=i).value for i in ra]
    aa, mm, uu, ii = [], [], [], []
    for row in WB.rows:
        udip = row[3]
        if udip.data_type == udip.TYPE_STRING and udip.value.find('-') > 0:
            m = int(udip.value.split('-')[0]) - 1
            for i, c in enumerate(row[5:]):
                if c.data_type == udip.TYPE_NUMERIC and c.value is not None:
                    # print(c.fill.fgColor.value)
                    if True:  # c.fill.bgColor.value == '00000000':  # black!
                        aa.append(A[i])
                        mm.append(m)
                        uu.append(c.value / 1000)
                        ii.append(current[i])
                elif c.data_type == c.TYPE_STRING and c.value.startswith('KE'):
                    # print('KE', c.coordinate)
                    cs = WS[c.coordinate]
                    if cs.data_type == udip.TYPE_NUMERIC:
                        # print(i, cs.value, cSmall[i])
                        if cs.value is not None and cSmall[i] is not None:
                            aa.append(A[i])
                            mm.append(m)
                            uu.append(cs.value / 1000)
                            ii.append(cSmall[i])
                    else:
                        raise Exception("no KE value on:" + str(c.coordinate))

    dataU = pg.DataContainerERT()
    dataU.setSensorPositions(pos)
    dataU.resize(len(aa))
    dataU.set('a', np.array(aa))
    dataU.set('b', np.array(aa)+1)
    dataU.set('m', np.array(mm)+1)
    dataU.set('n', np.array(mm))
    dataU.set('u', uu)
    dataU.set('i', ii)
    dataU.set('k', pb.geometricFactor(dataU))
    dataU.set('r', pg.abs(dataU('u')/dataU('i')))
    dataU.set('rhoa', pg.abs(dataU('r') * dataU('k')))
    dataU.markValid(dataU('u') > 0)
    dataU.checkDataValidity()
    dataU.removeInvalid()

    return dataU
